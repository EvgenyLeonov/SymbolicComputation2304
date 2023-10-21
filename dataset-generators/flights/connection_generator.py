import openpyxl
import random

wb = openpyxl.load_workbook("Topology.xlsx")
sheet_cities = wb["Cities"]
sheet_flights = wb["Flights"]

cities = []
edges = []

CITY_NAME_COLUMN_INDEX = 2

row_num = 2
need_exit = False

while not need_exit:
    cell = sheet_cities.cell(row=row_num, column=CITY_NAME_COLUMN_INDEX)
    val = cell.value
    if val is None:
        need_exit = True
    else:
        cities.append(val)
    row_num = row_num + 1

print("Cities:" + str(cities))
cities_count = len(cities)
print("count:" + str(cities_count))

manual_pairs = []
manual_pairs.append(["Poznan", "Krakov"])
manual_pairs.append(["Gdansk", "Warsaw"])
manual_pairs.append(["Gdansk", "Poznan"])
manual_pairs.append(["Krakov", "Warsaw"])
manual_pairs.append(["Dresden", "Warsaw"])
manual_pairs.append(["Berlin", "Warsaw"])
manual_pairs.append(["Prague", "Warsaw"])
manual_pairs.append(["Dresden", "Berlin"])
manual_pairs.append(["Prague", "Bratislava"])
manual_pairs.append(["Budapest", "Bratislava"])
manual_pairs.append(["Prague", "Brno"])
manual_pairs.append(["Budapest", "Belgrade"])
manual_pairs.append(["Tirane", "Belgrade"])
manual_pairs.append(["Rome", "Belgrade"])
manual_pairs.append(["Rijeka", "Zagreb"])
manual_pairs.append(["Budapest", "Zagreb"])
manual_pairs.append(["Graz", "Zagreb"])
manual_pairs.append(["Graz", "Salzburg"])
manual_pairs.append(["Vienna", "Salzburg"])
manual_pairs.append(["Innsbruck", "Salzburg"])
manual_pairs.append(["Magdeburg", "Berlin"])
manual_pairs.append(["Hamburg", "Berlin"])
manual_pairs.append(["Hamburg", "Frankfurt"])
manual_pairs.append(["Munich", "Frankfurt"])
manual_pairs.append(["Munich", "Verona"])
manual_pairs.append(["Rome", "Napoli"])
manual_pairs.append(["Tirane", "Napoli"])
manual_pairs.append(["Munich", "Innsbruck"])
manual_pairs.append(["Rome", "Verona"])
manual_pairs.append(["Napoli", "Rijeka"])
manual_pairs.append(["Tirane", "Rijeka"])
manual_pairs.append(["Prague", "Vienna"])
manual_pairs.append(["Prague", "Berlin"])
manual_pairs.append(["Dresden", "Magdeburg"])
manual_pairs.append(["Munich", "Magdeburg"])
manual_pairs.append(["Berlin", "Napoli"])
manual_pairs.append(["Warsaw", "Napoli"])
manual_pairs.append(["Bratislava", "Vienna"])
manual_pairs.append(["Zagreb", "Vienna"])
manual_pairs.append(["Prague", "Rome"])
manual_pairs.append(["Bratislava", "Rome"])
manual_pairs.append(["Warsaw", "Rome"])
manual_pairs.append(["Berlin", "Rome"])
manual_pairs.append(["Brno", "Graz"])
manual_pairs.append(["Munich", "Prague"])

for edge_pair in manual_pairs:
    price = random.randint(1, 2) * 100
    city1 = edge_pair[0]
    city2 = edge_pair[1]
    id1 = cities.index(city1) + 1
    id2 = cities.index(city2) + 1
    edges.append([city1, city2, price, id1, id2])

print("Flights:" + str(edges))
edges_count = len(edges)
print("count:" + str(edges_count))

# save this data on Flights sheet
row_num = 2
for edge in edges:
    sheet_flights.cell(row=row_num, column=1).value = row_num - 1
    col_ind = 2
    for c in edge:
        if col_ind > 4:
            break
        sheet_flights.cell(row=row_num, column=col_ind).value = c
        col_ind = col_ind + 1
    row_num = row_num + 1
wb.save("flights.xlsx")

# write data for viewer
lines = ["const imgFolder = 'img/';\r\n"]

# nodes
ind = 0
lines.append("const nodes = [\n")
for city in cities:
    city_id = cities.index(city) + 1
    lines.append("{ id: " + str(city_id) + ", shape: 'circularImage', image: imgFolder + '" + city + ".jpg', label: '" + city + "'  }")
    if ind < cities_count - 1:
        lines.append(",\n")
        ind = ind + 1
lines.append("];\r\n")

# edges
ind = 0
lines.append("const edges = [\n")
for edge in edges:
    price = edge[2]
    city1_id = edge[3]
    city2_id = edge[4]
    lines.append("{ from: " + str(city1_id) + ", to: " + str(city2_id) + ", label: '" + str(price) + "' }")
    if ind < edges_count - 1:
        lines.append(",\n")
        ind = ind + 1
lines.append("];\r\n")

with open("data.js", "w") as f:
    for line in lines:
        f.write(line)

# export to csv
with open("flights.csv", "w") as f:
    for edge in edges:
        for i in range(0, 3):
            f.write(str(edge[i]))
            f.write(",")
        f.write("\n")

